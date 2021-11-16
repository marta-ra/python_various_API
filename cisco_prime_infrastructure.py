import requests
import json
import datetime
from log_pass import LOGIN, PASSWORD
from openpyxl import load_workbook
import smtplib
import os


class sender_mail():
    SENDER = 'test@gmail.by'
    RECEIVERS_LIST = ['test1@gmail.by', 'test2@gmail.by']

    # function for sending email:
    def send_mail(self, points):
        points_inf = ''
        for point_data in points:
            points_inf += f'<tr><td>{point_data[0]}</td><td>{point_data[1]}</td><td>{point_data[2]}</td><td>{point_data[3]}</td><td>{point_data[4]}</td><td>{point_data[5]}</td><td>{point_data[6]}</td></tr>\n'
        for receiver_mail in self.RECEIVERS_LIST:
            # do not indent the text of "message", otherwise there will be empty email
            message = """From: From Person <test@gmail.by>
To: To Person <{0}>
MIME-Version: 1.0
Content-type: text/html
Subject: Report

<table border="1">
<caption>Point info</caption>
<tr>
<th>id</th>
<th>point in cisco-prime</th>
<th>point in file</th>
<th>name office</th>
<th>ip</th>
<th>Unavailable period</th>
<th>Shutdown time</th>
</tr>
{1}
</table>
""".format(receiver_mail, points_inf)
            print(message)
            # the sending of email is performed here:
            try:
                smtpObj = smtplib.SMTP('smtp-server.by', 25)
                smtpObj.sendmail(self.SENDER, receiver_mail, message)
                print('Success email')
            except smtplib.SMTPException:
                print('Fail email')


class NoDeviceFound(Exception):
    pass


class point():
    PI = 'cisco-prime-infrastructure'
    USER = LOGIN
    PASSWORD = PASSWORD
    BASE = 'https://%s:%s@%s/webacs/api/v4/' % (USER, PASSWORD, PI)
    path = r'\path'
    file_name = 'Data_points.xlsm'
    sheet_name = 'All_data'

    # Creation of values with which script will work further:
    def __init__(self):
        self.worksheet = self.sheet_active(self.path, self.file_name, self.sheet_name)
        # function "".max_row" gives the number of filled lines in the xlsx file:
        self.max_row_not_empty = self.worksheet.max_row
        self.now = datetime.datetime.now()
        self.all_AP_list = self.AccessPointDetails()

    # translation into datetime so that you can later compare/do mathematical operations with dates:
    def in_datetime(self, date_from_api):
        date_api = (date_from_api).replace('T', ' ').rpartition('.')[0]
        date_api = datetime.datetime.strptime(date_api, '%Y-%m-%d %H:%M:%S')
        return date_api

    # translation into convenient view for placing in email:
    def datetime_in_str(self, date_datetime):
        return date_datetime.strftime('%H:%M - %d/%m/%Y')

    # make the desired page active in xlsx file:
    def sheet_active(self, path, file_name, sheet_name):
        path = fr'{path}'
        workbook = load_workbook(os.path.join(path, file_name))
        for sheet in workbook:
            if sheet.title == sheet_name:
                workbook[sheet.title].views.sheetView[0].tabSelected = True
            else:
                workbook[sheet.title].views.sheetView[0].tabSelected = False
        workbook.active = workbook[sheet_name]
        return workbook.active

    # For API request for information on all points:
    def AccessPointDetails(self):
        # cisco gives only 1000 records in one request, so I make 2 requests with records up to 1000:
        result = requests.get(self.BASE + "data/AccessPointDetails.json?.full=true&.firstResult=0&.maxResults=1000",
                              verify=False)
        result_2 = requests.get(
            self.BASE + "data/AccessPointDetails.json?.full=true&.firstResult=1000&.maxResults=1000",
            verify=False)
        result.raise_for_status()
        # combine 2 results into one final list of points:
        all_AP_list = result.json()['queryResponse']['entity'] + result_2.json()['queryResponse']['entity']
        return all_AP_list

    # Other functions for API cisco prime:
    def all_devices(self):
        all_devices = []
        result = requests.get(self.BASE + "data/Devices.json?.full=true", verify=False)
        result.raise_for_status()
        for device in result.json()['queryResponse']['entity']:
            all_devices.append(device['devicesDTO']['@id'], device['devicesDTO']['ipAddress'])
        return all_devices

    def all_WlanProfiles(self):
        all_WlanProfiles = []
        result = requests.get(self.BASE + "data/WlanProfiles.json?.full=true", verify=False)
        result.raise_for_status()
        for device in result.json()['queryResponse']['entity']:
            all_WlanProfiles.append(device['wlanProfilesDTO']['@id'], device['wlanProfilesDTO']['@displayName'],
                  device['wlanProfilesDTO']['ssid'])
        return all_WlanProfiles


    # main function - search and output of a list of inaccessible points according to specified conditions:
    def run_find_unreachable(self):
        unreachable_points = []
        for AP in self.all_AP_list:
            if 'UNREACHABLE' in AP['accessPointDetailsDTO']['reachabilityStatus']:
                id_AP = AP['accessPointDetailsDTO']['@id']
                name_AP = AP['accessPointDetailsDTO']['name']
                dissociatedTime_AP = self.in_datetime(
                    AP['accessPointDetailsDTO']['unifiedApInfo']['lastDissociatedTime'])
                unavailability_time = self.now - dissociatedTime_AP
                if datetime.timedelta(hours=1) < unavailability_time < datetime.timedelta(hours=3):
                    for rownum in range(4, self.max_row_not_empty + 1):
                        name_AP_in_file = self.worksheet.cell(row=rownum, column=1).value
                        name_oficce = self.worksheet.cell(row=rownum, column=14).value
                        ip = self.worksheet.cell(row=rownum, column=2).value
                        if name_AP_in_file in name_AP:
                            time_disconnect = self.datetime_in_str(dissociatedTime_AP)
                            point_disconnect = id_AP, name_AP, name_AP_in_file, name_oficce, ip, \
                                               str(unavailability_time).partition('.')[0], time_disconnect
                            unreachable_points.append(point_disconnect)
                            break

        return unreachable_points


if __name__ == "__main__":
    obj_point = point()
    unreach_points = obj_point.run_find_unreachable()
    print(unreach_points)
    email = sender_mail()
    email_send = email.send_mail(unreach_points)

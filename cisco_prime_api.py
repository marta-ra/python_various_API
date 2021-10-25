import requests
import json
import datetime
from log_pass import LOGIN, PASSWORD
from openpyxl import load_workbook
import os


PI = 'server'
USER = LOGIN
PASSWORD = PASSWORD
BASE = "https://%s:%s@%s/webacs/api/v4/" %(USER, PASSWORD, PI)
path = r'\\location'
file_name = "data_Wifi.xlsm"
sheet_name = 'Wifi'
time_zone_correction = datetime.timedelta(hours=3)
requests.packages.urllib3.disable_warnings()


class NoDeviceFound(Exception):
    pass


def in_datetime(date_from_api):
    # formatting in type "datetime":
    date_api = (date_from_api).replace('T', ' ').rpartition('.')[0]
    date_api = datetime.datetime.strptime(date_api, '%Y-%m-%d %H:%M:%S')
    return date_api


def sheet_active(path, file_name, sheet_name):
    path = fr'{path}'
    workbook = load_workbook(os.path.join(path, file_name))
    for sheet in workbook:
        if sheet.title == sheet_name:
            workbook[sheet.title].views.sheetView[0].tabSelected = True
        else:
            workbook[sheet.title].views.sheetView[0].tabSelected = False
    workbook.active = workbook[sheet_name]
    return workbook.active


def AccessPointDetails():
    result = requests.get(BASE + "data/AccessPointDetails.json?.full=true&.firstResult=0&.maxResults=1000", verify=False)
    result_2 = requests.get(BASE + "data/AccessPointDetails.json?.full=true&.firstResult=1000&.maxResults=1000",
                          verify=False)
    result.raise_for_status()
    all_AP_list = result.json()['queryResponse']['entity'] + result_2.json()['queryResponse']['entity']
    return all_AP_list


def all_devices():
    print("Getting all devices")
    result = requests.get(BASE + "data/Devices.json?.full=true", verify=False)
    result.raise_for_status()
    print(result.json())
    for device in result.json()['queryResponse']['entity']:
        print(device['devicesDTO']['@id'], device['devicesDTO']['ipAddress'])


def all_WlanProfiles():
    print("Getting all devices")
    print("{0:6s} {1:10s}".format("ID", "IP address"))
    result = requests.get(BASE + "data/WlanProfiles.json?.full=true", verify=False)
    result.raise_for_status()
    print(result.json())
    for device in result.json()['queryResponse']['entity']:
        print(device['wlanProfilesDTO']['@id'], device['wlanProfilesDTO']['@displayName'], device['wlanProfilesDTO']['ssid'])


def device_by_id(id):
    print("Getting a specific device")
    result = requests.get(BASE + "data/Devices/%s.json?.full=true" % id, verify=False)
    result.raise_for_status()
    print(json.dumps(result.json(), indent=2))


def device_by_ip(ip):
    result = requests.get(BASE + "data/Devices.json?.full=true&ipAddress=%s" % ip, verify=False)
    print(BASE + "data/Devices.json?.full=true&ipAddress=%s" % ip)
    result.raise_for_status()
    if result.json()['queryResponse']['@count'] == "1":
        return result.json()
    else:
        raise NoDeviceFound("No device with ip: %s" %ip)


if __name__ == "__main__":


# search for points that are inaccessible from an hour to 5 hours:
    now = datetime.datetime.now()
    all_AP_list = AccessPointDetails()
    for AP in all_AP_list:
        # search for unavailable AccessPoints:
        if 'UNREACHABLE' in AP['accessPointDetailsDTO']['reachabilityStatus']:
            id_AP = AP['accessPointDetailsDTO']['@id']
            name_AP = AP['accessPointDetailsDTO']['name']
            dissociatedTime_AP = in_datetime(AP['accessPointDetailsDTO']['unifiedApInfo']['lastDissociatedTime'])
            unavailability_time = now - dissociatedTime_AP - time_zone_correction     
            if datetime.timedelta(hours=1) < unavailability_time < datetime.timedelta(hours=5):
                # to search for ip point in the file:
                worksheet = sheet_active(path, file_name, sheet_name)
                # To dynamically find the last filled row:
                max_row_not_empty = worksheet.max_row
                for rownum in range(2, max_row_not_empty + 1):
                    name_AP_in_file = worksheet.cell(row=rownum, column=1).value
                    if name_AP_in_file == name_AP:
                        ip_AP = worksheet.cell(row=rownum, column=2).value
                        print(id_AP, name_AP_in_file, ip_AP, dissociatedTime_AP - time_zone_correction)
